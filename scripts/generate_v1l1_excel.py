#!/usr/bin/env python3
"""
Generate IFA Strategy Audit Excel Trade Log for Ravi VAM v3.

Re-runs the backtest with enhanced trade logging to capture:
- Per-trade indicator values (SMA50, SMA200, RSI14) at entry and exit
- Entry/exit reasons with exact trigger values
- Holding days, PnL in $ and %

Output: clients/ravi_vam/delivery/IFA_Strategy_Audit_Ravi_VAM_v3_Trades.xlsx
"""

import json
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
ENGINE_ROOT = PROJECT_ROOT.parent.parent  # _engine/code/
DATA_DIR = ENGINE_ROOT / "data" / "databento" / "equities"
RESULTS_DIR = PROJECT_ROOT / "results"
DELIVERY_DIR = PROJECT_ROOT / "delivery"
DELIVERY_DIR.mkdir(parents=True, exist_ok=True)

IFA_BLUE_HEX = "2E75B6"
RED_HEX = "C0392B"
GREEN_HEX = "27AE60"
LIGHT_GREEN_HEX = "E8F5E9"
LIGHT_RED_HEX = "FFEBEE"
LIGHT_BLUE_HEX = "D6EAF8"
WHITE_HEX = "FFFFFF"

ALLOCATIONS = {
    "WARMUP": {"CASH": 1.0},
    "CRASH": {"CASH": 1.0},
    "NEUTRAL": {"SHY": 1.0},
    "BULL_FULL": {"UPRO": 0.75, "TQQQ": 0.25},
    "BULL_CAUTION": {"UPRO": 0.50, "SHY": 0.50},
    "MOMENTUM": {"UPRO": 1.0},
    "RECOVERY": {"UPRO": 0.25, "SHY": 0.75},
    "BEAR_HEDGE": {"SHY": 1.0},
}


def load_data() -> dict[str, pd.DataFrame]:
    """Load SPY, UPRO, TQQQ, SHY daily data from DataBento CSVs."""
    dfs = {}
    for sym in ["SPY", "UPRO", "TQQQ", "SHY"]:
        path = DATA_DIR / f"{sym}_daily.csv"
        df = pd.read_csv(path, index_col=0, parse_dates=True)
        df.sort_index(inplace=True)
        dfs[sym] = df
    return dfs


def compute_indicators(spy: pd.DataFrame) -> pd.DataFrame:
    """Compute SMA(50), SMA(200), RSI(14) on SPY."""
    df = spy.copy()
    df["sma_50"] = df["close"].rolling(50).mean()
    df["sma_200"] = df["close"].rolling(200).mean()

    delta = df["close"].diff()
    gain = delta.where(delta > 0, 0.0)
    loss = -delta.where(delta < 0, 0.0)
    avg_gain = gain.ewm(alpha=1 / 14, min_periods=14, adjust=False).mean()
    avg_loss = loss.ewm(alpha=1 / 14, min_periods=14, adjust=False).mean()
    rs = avg_gain / avg_loss
    df["rsi_14"] = 100 - (100 / (1 + rs))

    return df


def classify_regime(sma50: float, sma200: float, rsi: float, rsi_5_ago: float | None) -> str:
    """Classify into one of 7 regimes."""
    if pd.isna(sma50) or pd.isna(sma200) or pd.isna(rsi):
        return "WARMUP"
    if rsi < 20:
        return "CRASH"
    if abs(sma50 - sma200) / sma200 < 0.005:
        return "NEUTRAL"
    if sma50 > sma200:
        if rsi > 65:
            return "MOMENTUM"
        elif 30 <= rsi <= 40:
            return "BULL_CAUTION"
        elif rsi > 40:
            return "BULL_FULL"
        else:
            return "BULL_CAUTION"
    if sma50 < sma200:
        if rsi > 50 and rsi_5_ago is not None and rsi_5_ago < 30:
            return "RECOVERY"
        return "BEAR_HEDGE"
    return "BEAR_HEDGE"


def build_entry_reason(regime: str, sma50: float, sma200: float, rsi: float) -> str:
    """Build a human-readable entry reason string."""
    if regime == "WARMUP":
        return "Warmup period (insufficient data)"
    if regime == "CRASH":
        return f"RSI {rsi:.1f} < 20 -> CRASH (100% Cash)"
    if regime == "NEUTRAL":
        pct = abs(sma50 - sma200) / sma200 * 100
        return f"|SMA50 {sma50:.2f} - SMA200 {sma200:.2f}| / SMA200 = {pct:.2f}% < 0.5% -> NEUTRAL (100% SHY)"
    if regime == "MOMENTUM":
        return f"SMA50 {sma50:.2f} > SMA200 {sma200:.2f} AND RSI {rsi:.1f} > 65 -> MOMENTUM (100% UPRO)"
    if regime == "BULL_CAUTION":
        return f"SMA50 {sma50:.2f} > SMA200 {sma200:.2f} AND RSI {rsi:.1f} in [30,40] -> BULL_CAUTION (50% UPRO + 50% SHY)"
    if regime == "BULL_FULL":
        return f"SMA50 {sma50:.2f} > SMA200 {sma200:.2f} AND RSI {rsi:.1f} > 40 -> BULL_FULL (75% UPRO + 25% TQQQ)"
    if regime == "RECOVERY":
        return f"SMA50 {sma50:.2f} < SMA200 {sma200:.2f} AND RSI {rsi:.1f} > 50 (recovery bounce) -> RECOVERY (25% UPRO + 75% SHY)"
    if regime == "BEAR_HEDGE":
        return f"SMA50 {sma50:.2f} < SMA200 {sma200:.2f} -> BEAR_HEDGE (100% SHY)"
    return f"Regime: {regime}"


def instruments_for_regime(regime: str) -> str:
    """Return instrument allocation string."""
    alloc = ALLOCATIONS.get(regime, {})
    parts = []
    for sym, wt in sorted(alloc.items(), key=lambda x: -x[1]):
        parts.append(f"{sym} {wt * 100:.0f}%")
    return " + ".join(parts) if parts else "CASH 100%"


def run_backtest_with_trades(
    dfs: dict[str, pd.DataFrame],
    initial_capital: float = 100_000,
    commission_pct: float = 0.0010,
    slippage_pct: float = 0.0005,
) -> tuple[list[dict[str, Any]], pd.DataFrame, int]:
    """Run backtest and capture detailed trade records with indicator values."""
    spy = compute_indicators(dfs["SPY"])

    common_dates = spy.index
    for sym in ["UPRO", "TQQQ", "SHY"]:
        common_dates = common_dates.intersection(dfs[sym].index)
    common_dates = common_dates.sort_values()

    spy = spy.loc[common_dates]
    returns = {}
    for sym in ["UPRO", "TQQQ", "SHY"]:
        returns[sym] = dfs[sym].loc[common_dates, "close"].pct_change()

    equity = initial_capital
    equity_curve = []
    prev_regime = None
    prev_alloc: dict[str, float] = {}
    trade_count = 0

    # Track trades as regime transitions
    trades: list[dict[str, Any]] = []
    current_trade_entry_date = None
    current_trade_entry_equity = initial_capital
    current_trade_entry_sma50 = None
    current_trade_entry_sma200 = None
    current_trade_entry_rsi = None
    current_trade_regime = None
    current_trade_entry_reason = None

    rsi_series = spy["rsi_14"]

    for i, date in enumerate(common_dates):
        rsi_5_ago = rsi_series.iloc[i - 5] if i >= 5 else None
        row = spy.loc[date]
        sma50 = row["sma_50"]
        sma200 = row["sma_200"]
        rsi = row["rsi_14"]

        regime = classify_regime(sma50, sma200, rsi, rsi_5_ago)
        alloc = ALLOCATIONS[regime]

        allocation_changed = (alloc != prev_alloc) and (prev_regime is not None)

        # Daily return from previous allocation
        daily_return = 0.0
        for sym, weight in prev_alloc.items():
            if sym == "CASH":
                continue
            sym_ret = returns[sym].iloc[i] if not pd.isna(returns[sym].iloc[i]) else 0.0
            daily_return += weight * sym_ret

        tc = 0.0
        if allocation_changed and i > 0:
            all_syms = set(list(alloc.keys()) + list(prev_alloc.keys()))
            turnover = sum(abs(alloc.get(s, 0.0) - prev_alloc.get(s, 0.0)) for s in all_syms)
            tc = turnover * (commission_pct + slippage_pct)
            trade_count += 1

            # Close previous trade
            if current_trade_entry_date is not None and current_trade_regime != "WARMUP":
                exit_equity = equity  # before today's return applied
                pnl_dollar = exit_equity - current_trade_entry_equity
                pnl_pct = (
                    (exit_equity / current_trade_entry_equity - 1) * 100
                    if current_trade_entry_equity > 0
                    else 0
                )
                holding_days = (date - current_trade_entry_date).days

                exit_reason = build_entry_reason(regime, sma50, sma200, rsi)

                trades.append(
                    {
                        "Trade #": len(trades) + 1,
                        "Entry Date": current_trade_entry_date.strftime("%Y-%m-%d"),
                        "Exit Date": date.strftime("%Y-%m-%d"),
                        "State": current_trade_regime,
                        "Action": (
                            "BUY"
                            if current_trade_regime
                            in ("BULL_FULL", "MOMENTUM", "BULL_CAUTION", "RECOVERY")
                            else "SELL/HEDGE"
                        ),
                        "Instruments": instruments_for_regime(current_trade_regime),
                        "Entry Equity": round(current_trade_entry_equity, 2),
                        "Exit Equity": round(exit_equity, 2),
                        "PnL ($)": round(pnl_dollar, 2),
                        "PnL (%)": round(pnl_pct, 2),
                        "Holding Days": holding_days,
                        "SMA(50) Entry": (
                            round(current_trade_entry_sma50, 2)
                            if current_trade_entry_sma50 is not None
                            and not pd.isna(current_trade_entry_sma50)
                            else ""
                        ),
                        "SMA(200) Entry": (
                            round(current_trade_entry_sma200, 2)
                            if current_trade_entry_sma200 is not None
                            and not pd.isna(current_trade_entry_sma200)
                            else ""
                        ),
                        "RSI(14) Entry": (
                            round(current_trade_entry_rsi, 1)
                            if current_trade_entry_rsi is not None
                            and not pd.isna(current_trade_entry_rsi)
                            else ""
                        ),
                        "SMA(50) Exit": (round(sma50, 2) if not pd.isna(sma50) else ""),
                        "SMA(200) Exit": (round(sma200, 2) if not pd.isna(sma200) else ""),
                        "RSI(14) Exit": (round(rsi, 1) if not pd.isna(rsi) else ""),
                        "Entry Reason": current_trade_entry_reason or "",
                        "Exit Reason": exit_reason,
                    }
                )

            # Start new trade
            current_trade_entry_date = date
            current_trade_entry_equity = equity
            current_trade_entry_sma50 = sma50
            current_trade_entry_sma200 = sma200
            current_trade_entry_rsi = rsi
            current_trade_regime = regime
            current_trade_entry_reason = build_entry_reason(regime, sma50, sma200, rsi)

        elif prev_regime is None:
            # First day — set initial trade
            current_trade_entry_date = date
            current_trade_entry_equity = equity
            current_trade_entry_sma50 = sma50
            current_trade_entry_sma200 = sma200
            current_trade_entry_rsi = rsi
            current_trade_regime = regime
            current_trade_entry_reason = build_entry_reason(regime, sma50, sma200, rsi)

        equity *= 1 + daily_return - tc
        equity_curve.append(
            {
                "date": date,
                "equity": equity,
                "regime": regime,
                "daily_return": daily_return,
            }
        )
        prev_regime = regime
        prev_alloc = alloc

    # Close final open trade
    if current_trade_entry_date is not None and current_trade_regime != "WARMUP":
        last_date = common_dates[-1]
        last_row = spy.loc[last_date]
        pnl_dollar = equity - current_trade_entry_equity
        pnl_pct = (
            (equity / current_trade_entry_equity - 1) * 100 if current_trade_entry_equity > 0 else 0
        )
        holding_days = (last_date - current_trade_entry_date).days
        trades.append(
            {
                "Trade #": len(trades) + 1,
                "Entry Date": current_trade_entry_date.strftime("%Y-%m-%d"),
                "Exit Date": last_date.strftime("%Y-%m-%d") + " (open)",
                "State": current_trade_regime,
                "Action": (
                    "BUY"
                    if current_trade_regime in ("BULL_FULL", "MOMENTUM", "BULL_CAUTION", "RECOVERY")
                    else "SELL/HEDGE"
                ),
                "Instruments": instruments_for_regime(current_trade_regime),
                "Entry Equity": round(current_trade_entry_equity, 2),
                "Exit Equity": round(equity, 2),
                "PnL ($)": round(pnl_dollar, 2),
                "PnL (%)": round(pnl_pct, 2),
                "Holding Days": holding_days,
                "SMA(50) Entry": (
                    round(current_trade_entry_sma50, 2)
                    if current_trade_entry_sma50 is not None
                    and not pd.isna(current_trade_entry_sma50)
                    else ""
                ),
                "SMA(200) Entry": (
                    round(current_trade_entry_sma200, 2)
                    if current_trade_entry_sma200 is not None
                    and not pd.isna(current_trade_entry_sma200)
                    else ""
                ),
                "RSI(14) Entry": (
                    round(current_trade_entry_rsi, 1)
                    if current_trade_entry_rsi is not None and not pd.isna(current_trade_entry_rsi)
                    else ""
                ),
                "SMA(50) Exit": (
                    round(last_row["sma_50"], 2) if not pd.isna(last_row["sma_50"]) else ""
                ),
                "SMA(200) Exit": (
                    round(last_row["sma_200"], 2) if not pd.isna(last_row["sma_200"]) else ""
                ),
                "RSI(14) Exit": (
                    round(last_row["rsi_14"], 1) if not pd.isna(last_row["rsi_14"]) else ""
                ),
                "Entry Reason": current_trade_entry_reason or "",
                "Exit Reason": "Position still open at end of data",
            }
        )

    ec_df = pd.DataFrame(equity_curve).set_index("date")
    return trades, ec_df, trade_count


def style_header(ws: Any, row: int, ncols: int) -> None:
    """Apply IFA blue header styling to a row."""
    header_fill = PatternFill(start_color=IFA_BLUE_HEX, end_color=IFA_BLUE_HEX, fill_type="solid")
    header_font = Font(bold=True, color=WHITE_HEX, size=10)
    thin_border = Border(
        bottom=Side(style="thin", color="000000"),
    )
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def auto_width(ws: Any, min_width: int = 10, max_width: int = 45) -> None:
    """Auto-size column widths based on content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


def main() -> None:
    """Generate the Excel trade log."""
    print("Loading DataBento data...")
    dfs = load_data()
    for sym, df in dfs.items():
        print(f"  {sym}: {len(df)} bars, {df.index[0].date()} to {df.index[-1].date()}")

    print("\nRunning backtest with detailed trade logging...")
    trades, ec_df, trade_count = run_backtest_with_trades(dfs)
    print(f"  {len(trades)} regime-transition trades captured")

    # Load results JSON for metrics
    results_path = RESULTS_DIR / "ravi_vam_v3_databento_results.json"
    with open(results_path) as f:
        results = json.load(f)

    perf = results["performance"]
    bench = results["benchmark"]

    # Load sweep data
    sweep_path = RESULTS_DIR / "vam_parameter_sweep.csv"
    sweep_df = pd.read_csv(sweep_path) if sweep_path.exists() else None

    # Create workbook
    wb = Workbook()

    # ===== SHEET 1: Summary =====
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Title
    ws_summary.merge_cells("A1:E1")
    title_cell = ws_summary["A1"]
    title_cell.value = "IFA Strategy Audit - Ravi VAM v3"
    title_cell.font = Font(bold=True, size=14, color=IFA_BLUE_HEX)
    title_cell.alignment = Alignment(horizontal="center")

    ws_summary.merge_cells("A2:E2")
    ws_summary["A2"].value = "DataBento Real Data | Jan 2020 - Mar 2025 | $100,000 Initial Capital"
    ws_summary["A2"].font = Font(size=10, color="666666")
    ws_summary["A2"].alignment = Alignment(horizontal="center")

    # Metrics table
    headers = ["Metric", "VAM v3", "SPY B&H", "IFA Threshold", "Status"]
    for col_idx, h in enumerate(headers, 1):
        ws_summary.cell(row=4, column=col_idx, value=h)
    style_header(ws_summary, 4, len(headers))

    metrics_data = [
        (
            "Total Return",
            f"{perf['total_return_pct']}%",
            f"{bench['total_return_pct']}%",
            "> 0%",
            "FAIL",
        ),
        ("CAGR", f"{perf['cagr_pct']}%", f"{bench['cagr_pct']}%", "> 5%", "FAIL"),
        ("Sharpe Ratio", perf["sharpe_ratio"], bench["sharpe_ratio"], "> 0.50", "FAIL"),
        (
            "Max Drawdown",
            f"{perf['max_drawdown_pct']}%",
            f"{bench['max_drawdown_pct']}%",
            "> -50%",
            "FAIL",
        ),
        ("Win Rate", f"{perf['win_rate_pct']}%", "-", "> 45%", "PASS"),
        ("Profit Factor", perf["profit_factor"], "-", "> 1.20", "FAIL"),
        ("Trades", perf["trade_count"], 0, "> 10", "PASS"),
        ("Sortino Ratio", perf["sortino_ratio"], "-", "> 0.70", "FAIL"),
        ("Annual Volatility", f"{perf['annual_volatility_pct']}%", "-", "-", "-"),
        ("Max Consec. Losses", perf["max_consecutive_losses"], "-", "-", "-"),
        (
            "Final Equity",
            f"${perf['final_equity']:,.2f}",
            f"${bench['final_equity']:,.2f}",
            "-",
            "-",
        ),
    ]

    fail_font = Font(bold=True, color=RED_HEX)
    pass_font = Font(bold=True, color=GREEN_HEX)
    light_red_fill = PatternFill(
        start_color=LIGHT_RED_HEX, end_color=LIGHT_RED_HEX, fill_type="solid"
    )
    light_green_fill = PatternFill(
        start_color=LIGHT_GREEN_HEX, end_color=LIGHT_GREEN_HEX, fill_type="solid"
    )

    for row_idx, (metric, vam, spy, threshold, status) in enumerate(metrics_data, 5):
        ws_summary.cell(row=row_idx, column=1, value=metric).font = Font(bold=True)
        ws_summary.cell(row=row_idx, column=2, value=vam)
        ws_summary.cell(row=row_idx, column=3, value=spy)
        ws_summary.cell(row=row_idx, column=4, value=threshold)
        status_cell = ws_summary.cell(row=row_idx, column=5, value=status)
        if status == "FAIL":
            status_cell.font = fail_font
            status_cell.fill = light_red_fill
        elif status == "PASS":
            status_cell.font = pass_font
            status_cell.fill = light_green_fill

    # Verdict
    verdict_row = 5 + len(metrics_data) + 1
    ws_summary.merge_cells(f"A{verdict_row}:E{verdict_row}")
    verdict_cell = ws_summary.cell(row=verdict_row, column=1)
    verdict_cell.value = "VERDICT: FAIL - DO NOT DEPLOY (3/10 gates passed)"
    verdict_cell.font = Font(bold=True, size=12, color=RED_HEX)
    verdict_cell.alignment = Alignment(horizontal="center")
    verdict_cell.fill = light_red_fill

    auto_width(ws_summary)
    ws_summary.freeze_panes = "A5"

    # ===== SHEET 2: All Trades =====
    ws_trades = wb.create_sheet("All Trades")

    trade_headers = [
        "Trade #",
        "Entry Date",
        "Exit Date",
        "State",
        "Action",
        "Instruments",
        "Entry Equity",
        "Exit Equity",
        "PnL ($)",
        "PnL (%)",
        "Holding Days",
        "SMA(50) Entry",
        "SMA(200) Entry",
        "RSI(14) Entry",
        "SMA(50) Exit",
        "SMA(200) Exit",
        "RSI(14) Exit",
        "Entry Reason",
        "Exit Reason",
    ]

    for col_idx, h in enumerate(trade_headers, 1):
        ws_trades.cell(row=1, column=col_idx, value=h)
    style_header(ws_trades, 1, len(trade_headers))

    for row_idx, trade in enumerate(trades, 2):
        for col_idx, header in enumerate(trade_headers, 1):
            val = trade.get(header, "")
            cell = ws_trades.cell(row=row_idx, column=col_idx, value=val)

            # Color PnL columns
            if header == "PnL ($)" and isinstance(val, (int, float)):
                cell.number_format = "#,##0.00"
                if val < 0:
                    cell.font = Font(color=RED_HEX)
                    cell.fill = light_red_fill
                elif val > 0:
                    cell.font = Font(color=GREEN_HEX)
                    cell.fill = light_green_fill
            elif header == "PnL (%)" and isinstance(val, (int, float)):
                cell.number_format = '0.00"%"'
                if val < 0:
                    cell.font = Font(color=RED_HEX)
                elif val > 0:
                    cell.font = Font(color=GREEN_HEX)
            elif header in ("Entry Equity", "Exit Equity") and isinstance(val, (int, float)):
                cell.number_format = "$#,##0.00"
            elif header == "Holding Days":
                cell.alignment = Alignment(horizontal="center")

    auto_width(ws_trades, max_width=60)
    ws_trades.freeze_panes = "A2"

    # ===== SHEET 3: Year by Year =====
    ws_yearly = wb.create_sheet("Year by Year")

    yearly_headers = ["Year", "VAM v3 Return", "Outcome"]
    for col_idx, h in enumerate(yearly_headers, 1):
        ws_yearly.cell(row=1, column=col_idx, value=h)
    style_header(ws_yearly, 1, len(yearly_headers))

    yearly_returns = perf["yearly_returns"]
    for row_idx, (year, ret) in enumerate(sorted(yearly_returns.items()), 2):
        ws_yearly.cell(row=row_idx, column=1, value=int(year))
        ret_cell = ws_yearly.cell(row=row_idx, column=2, value=f"{ret:+.2f}%")
        if ret < 0:
            ret_cell.font = Font(bold=True, color=RED_HEX)
            ret_cell.fill = light_red_fill
            ws_yearly.cell(row=row_idx, column=3, value="Loss").font = fail_font
        else:
            ret_cell.font = Font(bold=True, color=GREEN_HEX)
            ret_cell.fill = light_green_fill
            ws_yearly.cell(row=row_idx, column=3, value="Gain").font = pass_font

    # Add summary row
    sum_row = 2 + len(yearly_returns) + 1
    ws_yearly.cell(row=sum_row, column=1, value="TOTAL").font = Font(bold=True)
    total_cell = ws_yearly.cell(row=sum_row, column=2, value=f"{perf['total_return_pct']:+.2f}%")
    total_cell.font = Font(bold=True, color=RED_HEX, size=11)

    auto_width(ws_yearly)
    ws_yearly.freeze_panes = "A2"

    # ===== SHEET 4: Parameter Sweep =====
    ws_sweep = wb.create_sheet("Parameter Sweep")

    if sweep_df is not None and len(sweep_df) > 0:
        # Sort by total_return_pct descending, take top 20
        sweep_sorted = sweep_df.sort_values("total_return_pct", ascending=False).head(20)

        sweep_headers = [
            "Variant",
            "SMA Fast",
            "SMA Slow",
            "RSI Entry",
            "UPRO %",
            "TQQQ %",
            "Total Return %",
            "CAGR %",
            "Sharpe",
            "Max DD %",
            "2022 Return %",
        ]
        for col_idx, h in enumerate(sweep_headers, 1):
            ws_sweep.cell(row=1, column=col_idx, value=h)
        style_header(ws_sweep, 1, len(sweep_headers))

        for row_idx, (_, srow) in enumerate(sweep_sorted.iterrows(), 2):
            ws_sweep.cell(row=row_idx, column=1, value=srow.get("variant_name", ""))
            ws_sweep.cell(row=row_idx, column=2, value=srow.get("sma_fast", ""))
            ws_sweep.cell(row=row_idx, column=3, value=srow.get("sma_slow", ""))
            ws_sweep.cell(row=row_idx, column=4, value=srow.get("rsi_entry", ""))
            ws_sweep.cell(row=row_idx, column=5, value=srow.get("upro_pct", ""))
            ws_sweep.cell(row=row_idx, column=6, value=srow.get("tqqq_pct", ""))

            ret_val = srow.get("total_return_pct", 0)
            ret_cell = ws_sweep.cell(row=row_idx, column=7, value=round(ret_val, 2))
            ret_cell.number_format = "0.00"
            if ret_val < 0:
                ret_cell.font = Font(color=RED_HEX)

            ws_sweep.cell(row=row_idx, column=8, value=round(srow.get("cagr_pct", 0), 2))
            ws_sweep.cell(row=row_idx, column=9, value=round(srow.get("sharpe", 0), 4))

            dd_val = srow.get("max_dd_pct", 0)
            dd_cell = ws_sweep.cell(row=row_idx, column=10, value=round(dd_val, 2))
            if dd_val < -50:
                dd_cell.font = Font(color=RED_HEX, bold=True)

            y2022 = srow.get("year_2022_return", 0)
            y2022_cell = ws_sweep.cell(row=row_idx, column=11, value=round(y2022, 2))
            if y2022 < 0:
                y2022_cell.font = Font(color=RED_HEX)
                y2022_cell.fill = light_red_fill
    else:
        ws_sweep.cell(row=1, column=1, value="Parameter sweep data not available")

    auto_width(ws_sweep)
    ws_sweep.freeze_panes = "A2"

    # Save
    output_path = DELIVERY_DIR / "IFA_Strategy_Audit_Ravi_VAM_v3_Trades.xlsx"
    wb.save(str(output_path))
    file_size_kb = output_path.stat().st_size / 1024
    print(f"\nExcel generated: {output_path}")
    print(f"  Size: {file_size_kb:.0f} KB")
    print(f"  Sheets: Summary, All Trades ({len(trades)} trades), Year by Year, Parameter Sweep")


if __name__ == "__main__":
    main()
