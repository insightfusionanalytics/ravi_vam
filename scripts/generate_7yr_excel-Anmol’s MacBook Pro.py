#!/usr/bin/env python3
"""
Generate professional Excel trade log for Ravi VAM 7.5yr strategy audit.

Reads:
  - results/databento_7yr_chart_data.json  (trades + equity curves for 7 strategies)
  - results/databento_7yr_384_sweep.csv    (full 384-strategy sweep results)

Outputs:
  - delivery/Ravi_VAM_7yr_TradeLog.xlsx with 5 sheets:
    1. Summary          — Ravi vs Top 5 comparison
    2. All 384 Strategies — full sweep, conditional formatting
    3. Ravi Original Trades — per-trade detail
    4. Best Strategy Trades — per-trade detail for #1 ranked
    5. Parameter Impact  — average Sharpe by parameter value
"""

import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent
RESULTS_DIR = PROJECT_ROOT / "results"
DELIVERY_DIR = PROJECT_ROOT / "delivery"
DELIVERY_DIR.mkdir(parents=True, exist_ok=True)

CHART_DATA_PATH = RESULTS_DIR / "databento_7yr_chart_data.json"
SWEEP_CSV_PATH = RESULTS_DIR / "databento_7yr_384_sweep.csv"
OUTPUT_PATH = DELIVERY_DIR / "Ravi_VAM_7yr_TradeLog.xlsx"

RAVI_STRATEGY = "SMA50_200_RSI45_UPRO75_SHY_delay0_cautionTrue_vix35"

# ---------------------------------------------------------------------------
# IFA brand colours
# ---------------------------------------------------------------------------
IFA_BLUE = "2E75B6"
DARK_BLUE = "1B4F72"
GREEN_HEX = "27AE60"
YELLOW_HEX = "F39C12"
RED_HEX = "C0392B"
LIGHT_GREEN = "E8F5E9"
LIGHT_YELLOW = "FFF9C4"
LIGHT_RED = "FFEBEE"
LIGHT_BLUE = "D6EAF8"
LIGHT_GREY = "F2F2F2"
WHITE = "FFFFFF"

# Reusable styles
HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=11)
HEADER_FILL = PatternFill(start_color=IFA_BLUE, end_color=IFA_BLUE, fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, color=DARK_BLUE, size=14)
SUBTITLE_FONT = Font(name="Calibri", italic=True, color="666666", size=10)
DATA_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def load_chart_data() -> dict:
    """Load chart data JSON with trades and equity curves."""
    with open(CHART_DATA_PATH) as f:
        return json.load(f)


def load_sweep_data() -> pd.DataFrame:
    """Load full 384-strategy sweep CSV."""
    return pd.read_csv(SWEEP_CSV_PATH)


def style_header_row(ws: any, row: int, max_col: int) -> None:
    """Apply IFA blue header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data_cell(
    cell: any,
    num_format: str | None = None,
    align: Alignment | None = None,
) -> None:
    """Apply standard data cell styling."""
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = align or CENTER
    if num_format:
        cell.number_format = num_format


def auto_width(ws: any, min_width: int = 10, max_width: int = 35) -> None:
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 3, max_width))


def apply_stripe(ws: any, start_row: int, end_row: int, max_col: int) -> None:
    """Apply alternating row shading."""
    stripe_fill = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
    for row in range(start_row, end_row + 1):
        if row % 2 == 0:
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).fill = stripe_fill


# =========================================================================
# Sheet 1: Summary
# =========================================================================
def build_summary_sheet(wb: Workbook, sweep_df: pd.DataFrame) -> None:
    """Build the Summary comparison sheet."""
    ws = wb.active
    ws.title = "Summary"

    # Title row
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "IFA Strategy Audit — Ravi VAM | DataBento 7.5yr (2018-2025)"
    title_cell.font = TITLE_FONT
    title_cell.alignment = LEFT

    # Data source note
    ws.merge_cells("A2:H2")
    note_cell = ws["A2"]
    note_cell.value = "Equities: DataBento | VIX: CBOE (canonical)"
    note_cell.font = SUBTITLE_FONT
    note_cell.alignment = LEFT

    # Blank row
    row = 4

    # Headers
    headers = [
        "Strategy",
        "Return %",
        "CAGR %",
        "MaxDD %",
        "Sharpe",
        "Calmar",
        "Win Rate %",
        "Trades",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header_row(ws, row, len(headers))

    # Get Ravi's strategy row
    ravi_row = sweep_df[sweep_df["strategy"] == RAVI_STRATEGY]
    top5 = sweep_df.head(5)

    # Combine: Ravi first (highlighted), then top 5
    row += 1
    ravi_highlight = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")

    def write_strategy_row(
        ws: any,
        row: int,
        r: pd.Series,
        label: str | None = None,
        highlight: PatternFill | None = None,
    ) -> None:
        """Write one strategy row with metrics."""
        name = label or r["strategy"]
        values = [
            name,
            r["total_return_pct"],
            r["cagr_pct"],
            r["max_drawdown_pct"],
            r["sharpe"],
            r["calmar"],
            r["win_rate_pct"],
            int(r["num_trades"]),
        ]
        formats = [
            None,
            "#,##0.00",
            "#,##0.00",
            "#,##0.00",
            "0.0000",
            "0.0000",
            "0.0",
            "#,##0",
        ]
        for col, (val, fmt) in enumerate(zip(values, formats), 1):
            cell = ws.cell(row=row, column=col, value=val)
            style_data_cell(cell, num_format=fmt, align=LEFT if col == 1 else CENTER)
            if highlight:
                cell.fill = highlight

    # Ravi's original
    if not ravi_row.empty:
        write_strategy_row(
            ws,
            row,
            ravi_row.iloc[0],
            label=f"RAVI ORIGINAL: {RAVI_STRATEGY}",
            highlight=ravi_highlight,
        )
        row += 1

    # Separator row
    ws.merge_cells(f"A{row}:H{row}")
    sep_cell = ws.cell(row=row, column=1, value="TOP 5 RECOMMENDED STRATEGIES")
    sep_cell.font = Font(name="Calibri", bold=True, color=DARK_BLUE, size=10)
    sep_cell.alignment = LEFT
    sep_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
    for c in range(1, len(headers) + 1):
        ws.cell(row=row, column=c).fill = sep_fill
    row += 1

    # Top 5
    for _, r in top5.iterrows():
        write_strategy_row(ws, row, r)
        row += 1

    # Rank note
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    rank_note = ws.cell(row=row, column=1)
    if not ravi_row.empty:
        ravi_rank = sweep_df["sharpe"].rank(ascending=False).loc[ravi_row.index[0]]
        rank_note.value = (
            f"Ravi's original strategy ranks #{int(ravi_rank)} of "
            f"{len(sweep_df)} tested (by Sharpe ratio)"
        )
    else:
        rank_note.value = "Ravi's strategy not found in sweep data."
    rank_note.font = Font(name="Calibri", italic=True, color="333333", size=10)

    auto_width(ws)


# =========================================================================
# Sheet 2: All 384 Strategies
# =========================================================================
def build_all_strategies_sheet(wb: Workbook, sweep_df: pd.DataFrame) -> None:
    """Build the full sweep sheet with conditional formatting."""
    ws = wb.create_sheet("All 384 Strategies")

    # Sort by Sharpe descending
    df = sweep_df.sort_values("sharpe", ascending=False).reset_index(drop=True)

    # Headers
    display_cols = [
        ("Rank", None),
        ("Strategy", None),
        ("SMA Fast", "#,##0"),
        ("SMA Slow", "#,##0"),
        ("RSI Thresh", "#,##0"),
        ("UPRO %", "#,##0"),
        ("Defensive", None),
        ("Delay", "#,##0"),
        ("Caution", None),
        ("VIX Kill", "0.0"),
        ("Return %", "#,##0.00"),
        ("CAGR %", "#,##0.00"),
        ("MaxDD %", "#,##0.00"),
        ("Sharpe", "0.0000"),
        ("Calmar", "0.0000"),
        ("Win Rate %", "0.0"),
        ("Trades", "#,##0"),
    ]
    for col, (h, _) in enumerate(display_cols, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(display_cols))

    # Conditional fill colours
    green_fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    yellow_fill = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")
    red_fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
    ravi_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")

    sharpe_col_idx = 14  # column N = Sharpe

    for i, (_, r) in enumerate(df.iterrows()):
        row = i + 2
        is_ravi = r["strategy"] == RAVI_STRATEGY

        row_data = [
            i + 1,
            r["strategy"],
            int(r["sma_fast"]),
            int(r["sma_slow"]),
            int(r["rsi_threshold"]),
            int(r["upro_pct"]),
            r["defensive"],
            int(r["sma_delay"]),
            str(r["bull_caution"]),
            r["vix_kill"] if pd.notna(r["vix_kill"]) else "None",
            r["total_return_pct"],
            r["cagr_pct"],
            r["max_drawdown_pct"],
            r["sharpe"],
            r["calmar"],
            r["win_rate_pct"],
            int(r["num_trades"]),
        ]

        for col, (val, (_, fmt)) in enumerate(zip(row_data, display_cols), 1):
            cell = ws.cell(row=row, column=col, value=val)
            style_data_cell(
                cell,
                num_format=fmt,
                align=LEFT if col == 2 else CENTER,
            )

        # Conditional formatting on Sharpe
        sharpe_val = r["sharpe"]
        if sharpe_val >= 1.5:
            fill = green_fill
        elif sharpe_val >= 1.0:
            fill = yellow_fill
        else:
            fill = red_fill

        # Highlight entire Ravi row in blue
        if is_ravi:
            fill = ravi_fill

        for col in range(1, len(display_cols) + 1):
            ws.cell(row=row, column=col).fill = fill

    # Freeze panes
    ws.freeze_panes = "A2"
    auto_width(ws, min_width=8, max_width=45)


# =========================================================================
# Sheet 3 & 4: Trade Detail
# =========================================================================
def build_trade_sheet(
    wb: Workbook,
    sheet_name: str,
    trades: list[dict],
    strategy_name: str,
) -> None:
    """Build a per-trade detail sheet from chart data trades."""
    ws = wb.create_sheet(sheet_name)

    # Title
    ws.merge_cells("A1:K1")
    title = ws["A1"]
    title.value = f"Trade Log: {strategy_name}"
    title.font = Font(name="Calibri", bold=True, color=DARK_BLUE, size=12)
    title.alignment = LEFT

    ws.merge_cells("A2:K2")
    ws["A2"].value = f"Total trades: {len(trades)}"
    ws["A2"].font = SUBTITLE_FONT

    # Headers
    header_row = 4
    headers = [
        "#",
        "Date",
        "Action",
        "From State",
        "To State",
        "SPY Price",
        "SMA Fast",
        "SMA Slow",
        "RSI",
        "VIX",
        "Portfolio Equity",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=col, value=h)
    style_header_row(ws, header_row, len(headers))

    # Compute PnL between trades
    buy_fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    sell_fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")

    # Determine action: moving TO a bull state = BUY, moving TO defensive = SELL
    bull_states = {"BULL_FULL", "BULL_CAUTION", "MOMENTUM", "RECOVERY"}

    for i, t in enumerate(trades):
        row = header_row + 1 + i
        to_state = t.get("to", "")
        from_state = t.get("from", "")
        action = "BUY" if to_state in bull_states else "SELL"

        row_data = [
            i + 1,
            t.get("date", ""),
            action,
            from_state,
            to_state,
            t.get("spy_close"),
            t.get("sma_fast"),
            t.get("sma_slow"),
            t.get("rsi"),
            t.get("vix"),
            t.get("equity"),
        ]

        formats = [
            "#,##0",
            None,
            None,
            None,
            None,
            "#,##0.00",
            "#,##0.00",
            "#,##0.00",
            "0.00",
            "0.00",
            "#,##0.00",
        ]

        for col, (val, fmt) in enumerate(zip(row_data, formats), 1):
            cell = ws.cell(row=row, column=col, value=val)
            style_data_cell(
                cell,
                num_format=fmt,
                align=LEFT if col in (2, 3, 4, 5) else CENTER,
            )

        # Colour the action column
        action_cell = ws.cell(row=row, column=3)
        action_cell.fill = buy_fill if action == "BUY" else sell_fill
        action_cell.font = Font(
            name="Calibri",
            bold=True,
            size=10,
            color=GREEN_HEX if action == "BUY" else RED_HEX,
        )

    # Stripe rows
    apply_stripe(
        ws,
        header_row + 1,
        header_row + len(trades),
        len(headers),
    )

    ws.freeze_panes = f"A{header_row + 1}"
    auto_width(ws)


# =========================================================================
# Sheet 5: Parameter Impact
# =========================================================================
def build_parameter_impact_sheet(wb: Workbook, sweep_df: pd.DataFrame) -> None:
    """Build parameter impact analysis sheet."""
    ws = wb.create_sheet("Parameter Impact")

    # Title
    ws.merge_cells("A1:D1")
    title = ws["A1"]
    title.value = "Parameter Impact Analysis — Average Sharpe by Parameter Value"
    title.font = Font(name="Calibri", bold=True, color=DARK_BLUE, size=12)
    title.alignment = LEFT

    row = 3
    headers = ["Parameter", "Value", "Avg Sharpe", "Count"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header_row(ws, row, len(headers))

    row += 1

    # Define parameter groups
    param_groups: list[tuple[str, str, list]] = [
        ("SMA Pair", "sma_fast", [50, 100]),
        ("RSI Threshold", "rsi_threshold", [35, 40, 45, 50]),
        ("UPRO Allocation", "upro_pct", [75, 100]),
        ("Defensive Asset", "defensive", ["SHY", "TLT", "cash"]),
    ]

    # VIX Kill: compare vix_kill == 35 vs NaN
    # Bull Caution: True vs False
    section_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
    best_fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")

    def write_group(
        ws: any,
        row: int,
        group_name: str,
        results: list[tuple[str, float, int]],
    ) -> int:
        """Write a parameter group and return next row."""
        # Section header
        ws.merge_cells(f"A{row}:D{row}")
        sec = ws.cell(row=row, column=1, value=group_name)
        sec.font = Font(name="Calibri", bold=True, color=DARK_BLUE, size=10)
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = section_fill
        row += 1

        # Find best
        best_sharpe = max(r[1] for r in results) if results else 0

        for label, avg_sharpe, count in results:
            vals = [None, str(label), avg_sharpe, count]
            fmts = [None, None, "0.0000", "#,##0"]
            for col, (val, fmt) in enumerate(zip(vals, fmts), 1):
                cell = ws.cell(row=row, column=col, value=val)
                style_data_cell(cell, num_format=fmt)
                if avg_sharpe == best_sharpe:
                    cell.fill = best_fill
            row += 1

        return row

    # Standard parameter groups
    for group_name, col_name, values in param_groups:
        results = []
        for v in values:
            subset = sweep_df[sweep_df[col_name] == v]
            if not subset.empty:
                results.append((str(v), round(subset["sharpe"].mean(), 4), len(subset)))
        row = write_group(ws, row, group_name, results)

    # VIX Kill
    vix_on = sweep_df[sweep_df["vix_kill"].notna()]
    vix_off = sweep_df[sweep_df["vix_kill"].isna()]
    results = []
    if not vix_on.empty:
        results.append(
            (
                "VIX Kill ON (35)",
                round(vix_on["sharpe"].mean(), 4),
                len(vix_on),
            )
        )
    if not vix_off.empty:
        results.append(
            (
                "VIX Kill OFF",
                round(vix_off["sharpe"].mean(), 4),
                len(vix_off),
            )
        )
    row = write_group(ws, row, "VIX Kill Switch", results)

    # Bull Caution
    caution_col = "bull_caution"
    results = []
    for val in [True, False]:
        subset = sweep_df[sweep_df[caution_col] == val]
        if not subset.empty:
            results.append(
                (
                    f"BULL_CAUTION = {val}",
                    round(subset["sharpe"].mean(), 4),
                    len(subset),
                )
            )
    row = write_group(ws, row, "Bull Caution Mode", results)

    # SMA Delay
    results = []
    for val in sorted(sweep_df["sma_delay"].unique()):
        subset = sweep_df[sweep_df["sma_delay"] == val]
        if not subset.empty:
            results.append(
                (
                    f"Delay = {int(val)}",
                    round(subset["sharpe"].mean(), 4),
                    len(subset),
                )
            )
    row = write_group(ws, row, "SMA Delay (days)", results)

    auto_width(ws, min_width=12)


# =========================================================================
# Main
# =========================================================================
def main() -> None:
    """Generate the full Ravi VAM 7yr Excel trade log."""
    print("Loading data...")
    chart_data = load_chart_data()
    sweep_df = load_sweep_data()
    sweep_df = sweep_df.sort_values("sharpe", ascending=False).reset_index(drop=True)

    print(f"  Sweep: {len(sweep_df)} strategies")
    print(f"  Chart data: {len(chart_data)} strategies with trades")

    # Identify best strategy (top Sharpe)
    best_strategy = sweep_df.iloc[0]["strategy"]
    print(f"  Best strategy: {best_strategy}")
    print(f"  Ravi strategy: {RAVI_STRATEGY}")

    # Check Ravi rank
    ravi_match = sweep_df[sweep_df["strategy"] == RAVI_STRATEGY]
    if not ravi_match.empty:
        ravi_rank = ravi_match.index[0] + 1
        print(f"  Ravi rank: #{ravi_rank}/{len(sweep_df)}")

    wb = Workbook()

    # Sheet 1: Summary
    print("Building Summary sheet...")
    build_summary_sheet(wb, sweep_df)

    # Sheet 2: All 384 Strategies
    print("Building All 384 Strategies sheet...")
    build_all_strategies_sheet(wb, sweep_df)

    # Sheet 3: Ravi Original Trades
    print("Building Ravi Original Trades sheet...")
    ravi_key = RAVI_STRATEGY
    if ravi_key in chart_data:
        build_trade_sheet(
            wb,
            "Ravi Original Trades",
            chart_data[ravi_key]["trades"],
            RAVI_STRATEGY,
        )
    else:
        # Try without vix suffix variations
        ravi_found = False
        for k in chart_data:
            if "RSI45_UPRO75_SHY" in k and "delay0_cautionTrue" in k:
                build_trade_sheet(
                    wb,
                    "Ravi Original Trades",
                    chart_data[k]["trades"],
                    k,
                )
                ravi_found = True
                print(f"  Matched Ravi to chart key: {k}")
                break
        if not ravi_found:
            ws = wb.create_sheet("Ravi Original Trades")
            ws["A1"].value = "Ravi strategy trades not found in chart data."
            print("  WARNING: Ravi strategy not found in chart data!")

    # Sheet 4: Best Strategy Trades
    print("Building Best Strategy Trades sheet...")
    if best_strategy in chart_data:
        build_trade_sheet(
            wb,
            "Best Strategy Trades",
            chart_data[best_strategy]["trades"],
            best_strategy,
        )
    else:
        # Best strategy might not be in chart data (only top 7 are)
        # Find the best strategy that IS in chart data
        best_in_chart = None
        for _, r in sweep_df.iterrows():
            if r["strategy"] in chart_data:
                best_in_chart = r["strategy"]
                break
        if best_in_chart:
            build_trade_sheet(
                wb,
                "Best Strategy Trades",
                chart_data[best_in_chart]["trades"],
                best_in_chart,
            )
            print(f"  Used best available in chart data: {best_in_chart}")
        else:
            ws = wb.create_sheet("Best Strategy Trades")
            ws["A1"].value = "Best strategy trades not found in chart data."
            print("  WARNING: No matching strategy in chart data!")

    # Sheet 5: Parameter Impact
    print("Building Parameter Impact sheet...")
    build_parameter_impact_sheet(wb, sweep_df)

    # Save
    wb.save(str(OUTPUT_PATH))
    print(f"\nSaved: {OUTPUT_PATH}")
    print("Done.")


if __name__ == "__main__":
    main()
